import openpyxl as p
import project 
import os
##tk open
def newuser():#creating a new user
    n=1
    while n!=0:
        a=input('Create your name and ur password:').split(',')
        for i in range(2,wa.max_row+1):
            if a[0]==wa.cell(row=i,column=1).value:print('\n\n!!!Username already exist!!!\n      TRY AGAIN\n');break
        else:
            n=0;a=list(a);a.extend([1,0]);wa.append(a)
    ask1='NO'
    return ask1
def exuser():#checking for existing user
    n=1
    while n!=0:
        a,b=input('Enter your name and ur password to login :').split(',')
        for i in range(2,wa.max_row+1):
            if a==wa.cell(row=i,column=1).value and b==wa.cell(row=i,column=2).value:
                print('Username and password is correct!!!!\n\n');n=0;break
            elif a==wa.cell(row=i,column=1).value and b!=wa.cell(row=i,column=2).value:
                print('\n\n!!!!!Password incorrect!!!!!\n      TRY AGAIN\n');break
        else:print('\n!User name not found!\n      TRY AGAIN\n')
    return(i)

def check():
    print(wa[chr(65)+str(v)].value,wa[chr(67)+str(v)].value,wa[chr(68)+str(v)].value)
    ask()
def reset():
    wa[chr(67)+str(v)].value,wa[chr(68)+str(v)].value=1,0
    ask()
def delete():
    wa.delete_rows(v)
    ask()

def ask():
    d=input('''if u want to check ur score press "u" ,"d" to delete ur account,"r" to reset ur acc''')
    if d in ('uUY'):
        check()
    elif d in 'DELETEdelete':
        delete()
    elif d in 'resetRESET':
        reset()
    else:
        print('session finished ')
absolutepath = os.path.abspath(__file__)
parentDirectory = os.path.dirname(absolutepath)
path= os.path.join(parentDirectory, 't.xlsx') 
wb=p.load_workbook(path)
wa=wb['Sheet1']
ask1=input("Are u a new user type 'yes' for new user or type 'no' for existing user:")#asking for user exist or not
if ask1 in'YESyes':
    print('This is a water sort game with infinite levels and u can also compare players and have separete user acc')
    ask1=newuser()
if ask1 in 'ONonNOno' :
    n='continue'
    v=exuser()
    print('WELCOME TO WATERSORT GAME')##game instructions and welcoming
    print('Rules:\n1.type tube"NO"-tube"NO"\n2.if each row has same elements then u win\n3.0 denotes empty\n\n')
    while n in 'continue':
        dh=2;dn=3
        lvl=wa.cell(row=v,column=3).value
        for i in range(1,int(lvl)):
            if i%2==0:
                dn+=1
            else:
                dh+=1
        print('LEVEl:',lvl)
        a=project.game(dh,dn)
        if a==0:#level increasing or not and also score 
            wa.cell(row=v,column=3).value=int(wa.cell(row=v,column=3).value)+1
            score=lvl*100
            wa.cell(row=v,column=4).value=int(wa.cell(row=v,column=4).value)+score
        else:
            score=lvl*10
            wa.cell(row=v,column=4).value=int(wa.cell(row=v,column=4).value)-score
        n=input('type "q" to quit or "c" to continue:')
        wb.save('D:\\programstuff\\progs\\pyProgs\\t.xlsx')
    ask()
    wb.save('D:\\programstuff\\progs\\pyProgs\\t.xlsx')