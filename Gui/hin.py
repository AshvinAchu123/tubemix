import os
import tkinter as t
from tkinter import *
from PIL import Image, ImageTk
import openpyxl as p

def clear_frame():
    for widget in window.winfo_children():widget.destroy()
def inpnew():
    n,b=user.get(),passw.get();a=(n,b)
    for i in range(2,wa.max_row+1):
        if a[0]==wa.cell(row=i,column=1).value:
            Label(text='!!!Username already exist!!!').place(x=230,y=500);break
    else:n=0;a=list(a);a.extend([1,0]);wa.append(a);screen_one();
def screen_two():
    clear_frame() 
    Label(window,text='username :').place(x=200,y=400);Entry(window,textvariable=user).place(x=300,y=400)
    Label(window,text='password :').place(x=200,y=430);Entry(window,textvariable=passw,show='•').place(x=300,y=430)
    Button(window,text='submit',command=inpnew).place(x=300,y=460)
def inp():
    a=user.get()
    b=passw.get()
    for i in range(2,wa.max_row+1):
        if a==wa.cell(row=i,column=1).value and b==wa.cell(row=i,column=2).value:
            Label(text='Username and password is correct!!!!').place(x=210,y=490);n=0;break
        elif a==wa.cell(row=i,column=1).value and b!=wa.cell(row=i,column=2).value:
            Label(text='         !!!!!Password incorrect!!!!!          ').place(x=210,y=490);break
    else:Label(text='      !User name not found!        ').place(x=210,y=490)
    return(i)
def screen_one():
    clear_frame()
    image1=Image.open("C:\\Users\\ashvi\\Desktop\\hh.png")
    test = ImageTk.PhotoImage(image1)
    label1 = t.Label(image=test)
    label1.image = test
    label1.place(x=260, y=200)
    Label(window,text='username :').place(x=200,y=400);Entry(window,textvariable=user).place(x=300,y=400)
    Label(window,text='password :').place(x=200,y=430);Entry(window,textvariable=passw,show='•').place(x=300,y=430)
    Button(window,text='submit',command=inp).place(x=300,y=460)
    Label(text='New to water sort ?').place(x=265,y=520)
    Button(window,text='Create account',command=screen_two).place(x=270,y=550)


window = t.Tk()
absolutepath = os.path.abspath(__file__)
parentDirectory = os.path.dirname(absolutepath)
path= os.path.join(parentDirectory, 't.xlsx') 
wb=p.load_workbook(path)
wa=wb['Sheet1']
window.title("Test")
window.geometry('700x900')
user = StringVar()
passw= StringVar()
screen_one()
window.mainloop()